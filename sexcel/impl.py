from openpyxl import Workbook, load_workbook
from io import BytesIO
import pathlib
import typing
from efc.interfaces.iopenpyxl import OpenpyxlInterface


class ExcelReader:
    def __init__(
        self,
        file: typing.Union[str, pathlib.Path, bytes, BytesIO],
        sheet: typing.Optional[str] = None,
    ):
        """
        :param file: name of a file or bytes
        :param sheet: name of the default sheet to access when a sheet is not specified
        """
        self._file: BytesIO
        if isinstance(file, bytes):
            self._file = BytesIO(file)
        elif isinstance(file, BytesIO):
            self._file = file
        else:
            path = file
            with open(path, "rb") as f:
                r = f.read()
                self._file = BytesIO(r)

        self._sheet: typing.Optional[str] = sheet

        self._wb: Workbook = load_workbook(self._file)
        self._interface: OpenpyxlInterface = OpenpyxlInterface(wb=self._wb, use_cache=True)

    def _calc_sheet(self, sheet: typing.Optional[str]) -> str:
        if sheet is None:
            # if sheet is not specified we read from the default or from the first in a file
            sheet = self._sheet or self._wb.sheetnames[0]
        return sheet

    def read_cell(
        self,
        column: str,
        row: int,
        *,
        calculate=False,
        sheet: typing.Optional[str] = None,
    ) -> typing.Any:
        """

        :param column: number of a column
        :param row: number of a row
        :param calculate: shall we calculate a formula?
        :param sheet: name of a sheet
        :return: value of a cell
        """

        sheet = self._calc_sheet(sheet)
        cell_name = f"{column}{row}"
        if calculate is True:
            return self._interface.calc_cell(cell_name, sheet)
        else:
            return self._wb[sheet][cell_name].value

    def read_cells(
        self,
        column: str,
        from_: int,
        to: int,
        sheet: typing.Optional[str] = None,
    ) -> typing.List[typing.Any]:
        """

        :param column: name of a column
        :param from_: from what a row from a column to read
        :param to: to what a row from a column to read
        :param sheet: name of a sheet
        :return: list of values of cells in the defined range
        """

        sheet = self._calc_sheet(sheet)
        return [v[0].value for v in self._wb[sheet][f"{column}{from_}":f"{column}{to}"]]

    def first_column_values(
        self, sheet: typing.Optional[str] = None
    ) -> typing.List[typing.Any]:
        """
        :param sheet: name of a sheet
        :return: list of values from the first column
        """

        sheet = self._calc_sheet(sheet)
        return [x.value for x in list(self._wb[sheet].columns)[0]]

    def last_column_values(
        self, sheet: typing.Optional[str] = None
    ) -> typing.List[typing.Any]:
        """
        :param sheet: name of a sheet
        :return: list of values from the last column
        """

        sheet = self._calc_sheet(sheet)
        return [x.value for x in list(self._wb[sheet].columns)[-1]]

    def sheet_size(self, sheet: typing.Optional[str] = None) -> typing.Tuple[int, int]:
        """

        :param sheet: name of a sheet
        :return: a tuple where the first element is a count of rows in a sheet and the second is a number of columns inthere
        """

        sheet = self._calc_sheet(sheet)
        return len(list(self._wb[sheet].rows)), len(list(self._wb[sheet].columns))
